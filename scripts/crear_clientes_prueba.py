"""
Crea un Excel de prueba con clientes nuevos y los sube a Odoo.
Uso: python scripts/crear_clientes_prueba.py
"""

import os
import sys
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------

API_URL = os.getenv("API_URL", "http://localhost:8000/odoo")
API_KEY = os.getenv("API_KEY", "")

if not API_KEY:
    env_path = os.path.join(os.path.dirname(__file__), "..", ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line.startswith("API_KEY="):
                    API_KEY = line.split("=", 1)[1]
                    break

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "..", "clientes_prueba.xlsx")

# ---------------------------------------------------------------------------
# Clientes de prueba
# ---------------------------------------------------------------------------

CLIENTES_NUEVOS = [
    {"nombre": "TechSmart Solutions SL",    "email": "contacto@techsmart.com",    "telefono": "+34 910 111 001", "nif": "B11111111"},
    {"nombre": "Innovatech Group SA",        "email": "info@innovatech.com",        "telefono": "+34 910 111 002", "nif": "A22222222"},
    {"nombre": "DataFlow Consulting SL",     "email": "hola@dataflow.com",          "telefono": "+34 910 111 003", "nif": "B33333333"},
    {"nombre": "CloudPeak Services SA",      "email": "ventas@cloudpeak.com",       "telefono": "+34 910 111 004", "nif": "A44444444"},
    {"nombre": "NexGen Digital SL",          "email": "contacto@nexgen.com",        "telefono": "+34 910 111 005", "nif": "B55555555"},
    {"nombre": "Bright Analytics SA",        "email": "info@brightanalytics.com",   "telefono": "+34 910 111 006", "nif": "A66666666"},
    {"nombre": "SmartBridge Corp SL",        "email": "hola@smartbridge.com",       "telefono": "+34 910 111 007", "nif": "B77777777"},
    {"nombre": "Vertex Systems SA",          "email": "ventas@vertexsys.com",       "telefono": "+34 910 111 008", "nif": "A88888888"},
    {"nombre": "PulseCode Solutions SL",     "email": "contacto@pulsecode.com",     "telefono": "+34 910 111 009", "nif": "B99999999"},
    {"nombre": "ElevateAI Technologies SA",  "email": "info@elevateai.com",         "telefono": "+34 910 111 010", "nif": "A10101010"},
    {"nombre": "Real Madrid CF",             "email": "info@realmadrid.com",        "telefono": "+34 910 111 011", "nif": "B10101011"},
    
    
]


# ---------------------------------------------------------------------------
# Crear Excel
# ---------------------------------------------------------------------------

def crear_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Nuevos_Contactos"

    # Estilos
    header_font  = Font(bold=True, size=11, color="FFFFFF")
    header_fill  = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    title_font   = Font(bold=True, size=13, color="375623")
    tip_font     = Font(italic=True, size=9, color="7F7F7F")
    row_fill_par = PatternFill(start_color="EBF3E8", end_color="EBF3E8", fill_type="solid")
    border       = Border(bottom=Side(style="thin", color="C6EFCE"))

    # Anchos de columna
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18

    # Fila 1: titulo
    ws.append(["CLIENTES DE PRUEBA - API-Odoo Middleware"])
    ws.merge_cells("A1:E1")
    ws.cell(1, 1).font = title_font
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    # Fila 2: instruccion
    ws.append(["La columna ID esta vacia: estos registros se CREARAN en Odoo al ejecutar excel_a_odoo.py"])
    ws.merge_cells("A2:E2")
    ws.cell(2, 1).font = tip_font

    # Fila 3: vacia
    ws.append([])

    # Fila 4: cabeceras
    ws.append(["ID (auto)", "Nombre *", "Email", "Telefono", "NIF/RUC"])
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Filas de datos
    for i, c in enumerate(CLIENTES_NUEVOS, 5):
        ws.append(["", c["nombre"], c["email"], c["telefono"], c["nif"]])
        fill = row_fill_par if i % 2 == 0 else None
        for cell in ws[ws.max_row]:
            if fill:
                cell.fill = fill
            cell.border = border

    # Filtros
    ws.auto_filter.ref = f"A4:E{ws.max_row}"

    wb.save(OUTPUT_FILE)
    print(f"Excel creado: {os.path.abspath(OUTPUT_FILE)}")
    print(f"Clientes listos para subir: {len(CLIENTES_NUEVOS)}")
    return OUTPUT_FILE


# ---------------------------------------------------------------------------
# Subir a Odoo
# ---------------------------------------------------------------------------

def enviar_a_odoo(model, method, args=None, kwargs=None):
    response = requests.post(
        API_URL,
        json={"model": model, "method": method, "args": args or [], "kwargs": kwargs or {}},
        headers={"X-Api-Key": API_KEY},
        timeout=30,
    )
    response.raise_for_status()
    return response.json()["result"]


def obtener_nifs_existentes():
    """Consulta Odoo y devuelve el conjunto de NIFs que ya existen."""
    nifs = [c["nif"] for c in CLIENTES_NUEVOS if c.get("nif")]
    if not nifs:
        return set()
    resultado = enviar_a_odoo(
        "res.partner",
        "search_read",
        args=[[["vat", "in", nifs]]],
        kwargs={"fields": ["vat"]},
    )
    return {r["vat"] for r in resultado if r.get("vat")}


def subir_clientes():
    print(f"\nVerificando cuales clientes ya existen en Odoo...")
    print(f"Middleware: {API_URL}")

    try:
        nifs_existentes = obtener_nifs_existentes()
    except Exception as e:
        print(f"  Error al consultar Odoo: {e}")
        nifs_existentes = set()

    pendientes = [c for c in CLIENTES_NUEVOS if c.get("nif") not in nifs_existentes]
    omitidos   = len(CLIENTES_NUEVOS) - len(pendientes)

    if omitidos:
        print(f"  {omitidos} cliente(s) ya existen en Odoo -> omitidos")

    if not pendientes:
        print("  Todos los clientes ya estan en Odoo. Nada que subir.")
        return 0, 0

    print(f"  {len(pendientes)} cliente(s) nuevos para crear\n")

    exitos, errores = 0, 0
    for i, c in enumerate(pendientes, 1):
        datos = {
            "name":          c["nombre"],
            "email":         c["email"],
            "phone":         c["telefono"],
            "vat":           c["nif"],
            "customer_rank": 1,
        }
        try:
            nuevo_id = enviar_a_odoo("res.partner", "create", args=[datos])
            print(f"  [{i:02d}] {c['nombre']:<40} -> creado OK (ID: {nuevo_id})")
            exitos += 1
        except Exception as e:
            print(f"  [{i:02d}] {c['nombre']:<40} -> ERROR: {e}")
            errores += 1

    print(f"\nResumen: {exitos} creados, {errores} errores, {omitidos} ya existian")
    return exitos, errores


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("  CREAR CLIENTES DE PRUEBA - API-Odoo")
    print("=" * 60)

    # Paso 1: crear el Excel
    crear_excel()

    # Paso 2: subir a Odoo
    print()
    respuesta = input("Subir los clientes a Odoo ahora? [s/N]: ").strip().lower()
    if respuesta in ("s", "si", "y", "yes"):
        subir_clientes()
        print("\nListo! Verifica los contactos en tu Odoo.")
    else:
        print(f"\nExcel guardado. Cuando quieras subirlos ejecuta:")
        print(f"  python scripts/excel_a_odoo.py clientes_prueba.xlsx")


if __name__ == "__main__":
    main()
