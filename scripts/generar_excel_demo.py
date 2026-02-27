"""
Genera un archivo Excel bidireccional para sincronizar con Odoo:
- Hoja "Contactos"        -> Datos de Odoo (editar aqui para actualizar en Odoo)
- Hoja "Nuevos_Contactos" -> Plantilla para agregar nuevos registros a Odoo
- Hoja "Instrucciones"    -> Guia de uso paso a paso
- Hoja "Configuracion"    -> Parametros del middleware

Flujo:
  Odoo -> Excel : python scripts/generar_excel_demo.py
  Excel -> Odoo : python scripts/excel_a_odoo.py

Uso: python scripts/generar_excel_demo.py
"""

import os
import sys
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------

API_URL = os.getenv("API_URL", "http://localhost:8000/odoo")
API_KEY = os.getenv("API_KEY", "")

# Si no hay API_KEY en env, intentar leer del .env del proyecto
if not API_KEY:
    env_path = os.path.join(os.path.dirname(__file__), "..", ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line.startswith("API_KEY="):
                    API_KEY = line.split("=", 1)[1]
                    break

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "..", "demo_odoo.xlsx")


def consultar_odoo(model, method, args=None, kwargs=None):
    """Consulta al middleware API-Odoo."""
    response = requests.post(
        API_URL,
        json={
            "model": model,
            "method": method,
            "args": args or [],
            "kwargs": kwargs or {},
        },
        headers={"X-Api-Key": API_KEY},
        timeout=30,
    )
    response.raise_for_status()
    data = response.json()
    if "result" not in data:
        raise ValueError(f"Respuesta inesperada: {data}")
    return data["result"]


def crear_hoja_config(wb):
    """Crea la hoja de configuracion."""
    ws = wb.create_sheet("Configuracion")

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60

    ws.append(["Parametro", "Valor"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws.append(["API URL", API_URL])
    ws.append(["API Key", API_KEY[:8] + "..." if API_KEY else "(no configurada)"])
    ws.append([""])
    ws.append(["Modelos disponibles", "res.partner, account.move, sale.order, purchase.order, stock.picking, product.template"])
    ws.append(["Metodos disponibles", "search_read, read, fields_get, name_search"])

    return ws


def crear_hoja_nuevos_contactos(wb):
    """Crea la hoja plantilla para agregar nuevos contactos a Odoo."""
    ws = wb.create_sheet("Nuevos_Contactos")

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    tip_font = Font(italic=True, size=9, color="7F7F7F")
    example_fill = PatternFill(start_color="EBF3E8", end_color="EBF3E8", fill_type="solid")
    border = Border(bottom=Side(style="thin", color="C6EFCE"))

    # Titulo
    ws.append(["NUEVOS CONTACTOS - Agrega aqui los registros que quieres crear en Odoo"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="375623")
    ws.merge_cells("A1:E1")
    ws.append(["Deja la columna ID vacia. Al enviar a Odoo se asignara automaticamente."])
    ws.cell(row=2, column=1).font = tip_font
    ws.merge_cells("A2:E2")
    ws.append([])

    # Cabeceras (fila 4)
    headers = ["ID (auto)", "Nombre *", "Email", "Telefono", "NIF/RUC"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Fila de ejemplo
    ws.append(["", "Nueva Empresa SL", "contacto@nuevaempresa.com", "+34 600 000 000", "B12345678"])
    for cell in ws[5]:
        cell.fill = example_fill
        cell.font = Font(italic=True, color="375623")
        cell.border = border

    # Filas vacias para rellenar (10 filas)
    for _ in range(10):
        ws.append(["", "", "", "", ""])
        for cell in ws[ws.max_row]:
            cell.border = border

    # Anchos de columna
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18

    ws.auto_filter.ref = f"A4:E{ws.max_row}"

    return ws


def crear_hoja_instrucciones(wb):
    """Crea la hoja de instrucciones."""
    ws = wb.create_sheet("Instrucciones")
    ws.column_dimensions["A"].width = 90

    title_font = Font(bold=True, size=14, color="2F5496")
    subtitle_font = Font(bold=True, size=11)
    cmd_font = Font(name="Courier New", size=10, color="C00000")

    instrucciones = [
        ("COMO USAR ESTE EXCEL CON EL MIDDLEWARE API-ODOO", title_font),
        ("", None),
        ("ODOO  ->  EXCEL  (descargar datos)", subtitle_font),
        ("   1. Asegurate de que el middleware este corriendo:", None),
        ("      uvicorn api:app --reload", cmd_font),
        ("   2. Ejecuta el script de generacion:", None),
        ("      python scripts/generar_excel_demo.py", cmd_font),
        ("   3. Abre el archivo demo_odoo.xlsx", None),
        ("   La hoja 'Contactos' se llena con los datos actuales de Odoo.", None),
        ("", None),
        ("EXCEL  ->  ODOO  (subir cambios y nuevos registros)", subtitle_font),
        ("   OPCION A: Actualizar contactos existentes", None),
        ("      - Edita cualquier fila en la hoja 'Contactos' (no borres el ID)", None),
        ("      - Guarda el archivo Excel", None),
        ("      - Ejecuta: python scripts/excel_a_odoo.py", cmd_font),
        ("      - Los registros con ID se actualizan en Odoo", None),
        ("", None),
        ("   OPCION B: Crear nuevos contactos", None),
        ("      - Ve a la hoja 'Nuevos_Contactos'", None),
        ("      - Rellena las filas (deja la columna ID vacia)", None),
        ("      - Guarda el archivo Excel", None),
        ("      - Ejecuta: python scripts/excel_a_odoo.py", cmd_font),
        ("      - Los registros sin ID se crean como nuevos en Odoo", None),
        ("", None),
        ("CAMPOS DISPONIBLES", subtitle_font),
        ("   Nombre (*obligatorio), Email, Telefono, NIF/RUC", None),
        ("", None),
        ("REQUISITOS", subtitle_font),
        ("   - Middleware corriendo: uvicorn api:app --reload", None),
        ("   - API Key correcta en el .env del proyecto", None),
    ]

    for texto, font in instrucciones:
        ws.append([texto])
        if font:
            ws.cell(row=ws.max_row, column=1).font = font

    return ws


def crear_hoja_contactos(wb, contactos):
    """Crea la hoja de contactos con datos de Odoo."""
    ws = wb.active
    ws.title = "Contactos"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    border = Border(
        bottom=Side(style="thin", color="D9E2F3"),
    )

    # Cabeceras
    headers = ["ID", "Nombre", "Email", "Telefono", "NIF/RUC"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Anchos de columna
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # Datos
    for c in contactos:
        ws.append([
            c.get("id", ""),
            c.get("name", ""),
            c.get("email", "") or "",
            c.get("phone", "") or "",
            c.get("vat", "") or "",
        ])
        for cell in ws[ws.max_row]:
            cell.border = border

    # Filtros
    ws.auto_filter.ref = f"A1:E{ws.max_row}"

    return ws


def main():
    print("Conectando al middleware API-Odoo...")
    print(f"  URL: {API_URL}")
    print(f"  API Key: {API_KEY[:8]}..." if API_KEY else "  API Key: (no configurada)")

    try:
        contactos = consultar_odoo(
            model="res.partner",
            method="search_read",
            args=[[["active", "=", True]]],
            kwargs={
                "fields": ["id", "name", "email", "phone", "vat"],
                "limit": 100,
                "order": "name asc",
            },
        )
        print(f"  Contactos obtenidos: {len(contactos)}")
    except Exception as e:
        print(f"  Error al consultar Odoo: {e}")
        print("  Generando Excel con datos de ejemplo...")
        contactos = [
            {"id": 1, "name": "Empresa Demo SA", "email": "info@demo.com", "phone": "+1234567890", "vat": "A12345678"},
            {"id": 2, "name": "Cliente Ejemplo SL", "email": "cliente@ejemplo.com", "phone": "+0987654321", "vat": "B87654321"},
            {"id": 3, "name": "Proveedor Test", "email": "proveedor@test.com", "phone": "+1122334455", "vat": "C11223344"},
        ]

    wb = Workbook()
    crear_hoja_contactos(wb, contactos)
    crear_hoja_nuevos_contactos(wb)
    crear_hoja_instrucciones(wb)
    crear_hoja_config(wb)

    wb.save(OUTPUT_FILE)
    print(f"\nExcel generado: {os.path.abspath(OUTPUT_FILE)}")
    print("Hojas disponibles:")
    print("  - Contactos        -> Edita y envia cambios a Odoo")
    print("  - Nuevos_Contactos -> Agrega aqui registros nuevos para crear en Odoo")
    print("  - Instrucciones    -> Guia de uso")
    print("  - Configuracion    -> Parametros del middleware")


if __name__ == "__main__":
    main()
