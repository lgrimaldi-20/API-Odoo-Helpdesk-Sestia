"""
Envia datos desde Excel al middleware API-Odoo.

Procesa dos hojas:
  - "Contactos"        -> Filas con ID = actualiza en Odoo (write)
  - "Nuevos_Contactos" -> Filas sin ID = crea en Odoo (create)

Uso: python scripts/excel_a_odoo.py [archivo.xlsx]
"""

import os
import sys
import requests
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------

API_URL = os.getenv("API_URL", "http://localhost:8000/odoo")
API_KEY = os.getenv("API_KEY", "")

if not API_KEY:
    env_path = os.path.join(os.path.dirname(__file__), "..", ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line.startswith("API_KEY="):
                    API_KEY = line.split("=", 1)[1]
                    break


# ---------------------------------------------------------------------------
# Comunicacion con el middleware
# ---------------------------------------------------------------------------

def enviar_a_odoo(model, method, args=None, kwargs=None):
    """Envia una peticion al middleware API-Odoo."""
    response = requests.post(
        API_URL,
        json={
            "model": model,
            "method": method,
            "args": args or [],
            "kwargs": kwargs or {},
        },
        headers={"X-Api-Key": API_KEY},
        timeout=30,
    )
    response.raise_for_status()
    data = response.json()
    if "error" in str(data):
        raise ValueError(data)
    return data["result"]


# ---------------------------------------------------------------------------
# Lectura de Excel
# ---------------------------------------------------------------------------

def leer_hoja(wb, nombre_hoja, fila_inicio=1):
    """
    Lee una hoja del workbook y devuelve lista de diccionarios.
    fila_inicio: fila donde estan las cabeceras (1-based).
    """
    if nombre_hoja not in wb.sheetnames:
        return []

    ws = wb[nombre_hoja]
    rows = list(ws.iter_rows(values_only=True))

    # Buscar la fila de cabeceras (la primera que tenga contenido util)
    header_row_idx = fila_inicio - 1
    for i, row in enumerate(rows):
        valores = [str(v).strip() for v in row if v is not None]
        if any(v for v in valores if v and not v.startswith("NUEVOS") and not v.startswith("Deja")):
            header_row_idx = i
            break

    if header_row_idx >= len(rows):
        return []

    headers = [str(h).lower().strip() if h else "" for h in rows[header_row_idx]]
    registros = []
    for row in rows[header_row_idx + 1:]:
        # Ignorar filas completamente vacias
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        registro = dict(zip(headers, row))
        registros.append(registro)

    return registros


def normalizar_registro(registro):
    """Extrae y normaliza los campos del registro independientemente del nombre de columna."""
    def get(*claves):
        for k in claves:
            v = registro.get(k)
            if v is not None and str(v).strip():
                return str(v).strip()
        return None

    odoo_id_raw = get("id", "id (auto)")
    try:
        odoo_id = int(float(odoo_id_raw)) if odoo_id_raw else None
    except (ValueError, TypeError):
        odoo_id = None

    datos = {
        "name":  get("nombre *", "nombre", "name"),
        "email": get("email"),
        "phone": get("telefono", "phone"),
        "vat":   get("nif/ruc", "vat"),
    }
    # Eliminar campos vacios
    datos = {k: v for k, v in datos.items() if v}

    return odoo_id, datos


# ---------------------------------------------------------------------------
# Sincronizacion
# ---------------------------------------------------------------------------

def procesar_actualizaciones(registros):
    """Procesa la hoja Contactos: actualiza registros con ID existente."""
    if not registros:
        return 0, 0

    print(f"\n[Contactos] Procesando {len(registros)} registros para actualizar...")
    exitos, errores = 0, 0

    for i, reg in enumerate(registros, 1):
        odoo_id, datos = normalizar_registro(reg)
        nombre = datos.get("name") or f"fila {i}"

        if not odoo_id:
            print(f"  [{i}] {nombre} -> omitido (sin ID, usa la hoja 'Nuevos_Contactos' para crear)")
            continue

        if not datos:
            print(f"  [{i}] ID={odoo_id} -> omitido (sin datos para actualizar)")
            continue

        try:
            enviar_a_odoo("res.partner", "write", args=[[odoo_id], datos])
            print(f"  [{i}] {nombre} (ID:{odoo_id}) -> actualizado OK")
            exitos += 1
        except Exception as e:
            print(f"  [{i}] {nombre} (ID:{odoo_id}) -> ERROR: {e}")
            errores += 1

    return exitos, errores


def procesar_nuevos(registros):
    """Procesa la hoja Nuevos_Contactos: crea registros sin ID."""
    if not registros:
        return 0, 0

    print(f"\n[Nuevos_Contactos] Procesando {len(registros)} registros para crear...")
    exitos, errores = 0, 0

    for i, reg in enumerate(registros, 1):
        odoo_id, datos = normalizar_registro(reg)

        if odoo_id:
            print(f"  [{i}] ID={odoo_id} -> omitido (ya tiene ID, muevelo a la hoja 'Contactos' para actualizar)")
            continue

        nombre = datos.get("name")
        if not nombre:
            print(f"  [{i}] -> omitido (nombre obligatorio)")
            continue

        try:
            nuevo_id = enviar_a_odoo("res.partner", "create", args=[datos])
            print(f"  [{i}] {nombre} -> creado OK (nuevo ID: {nuevo_id})")
            exitos += 1
        except Exception as e:
            print(f"  [{i}] {nombre} -> ERROR: {e}")
            errores += 1

    return exitos, errores


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    filepath = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(__file__), "..", "demo_odoo.xlsx"
    )

    if not os.path.exists(filepath):
        print(f"Archivo no encontrado: {filepath}")
        sys.exit(1)

    print(f"Archivo  : {os.path.abspath(filepath)}")
    print(f"Middleware: {API_URL}")
    print(f"API Key  : {API_KEY[:8]}..." if API_KEY else "API Key  : (no configurada)")

    wb = load_workbook(filepath, read_only=True)

    # Procesar hoja Contactos (actualizaciones)
    registros_existentes = leer_hoja(wb, "Contactos")
    exitos_act, errores_act = procesar_actualizaciones(registros_existentes)

    # Procesar hoja Nuevos_Contactos (creaciones)
    registros_nuevos = leer_hoja(wb, "Nuevos_Contactos", fila_inicio=4)
    exitos_new, errores_new = procesar_nuevos(registros_nuevos)

    wb.close()

    # Resumen final
    print("\n" + "=" * 50)
    print("RESUMEN")
    print(f"  Actualizados : {exitos_act} OK, {errores_act} errores")
    print(f"  Creados      : {exitos_new} OK, {errores_new} errores")
    print(f"  Total errores: {errores_act + errores_new}")
    print("=" * 50)


if __name__ == "__main__":
    main()
